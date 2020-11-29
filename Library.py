import os
import re
import subprocess
import sys

def RunCommand(Cmd, Input=None):
    RetValue = []
    Process = subprocess.Popen(Cmd, stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT, stdin=subprocess.PIPE)
    print(Process.communicate(Input)[0])

    return RetValue

class InfLibraryFile:
  RELibraryLine = re.compile(r"^\s*LIBRARY_CLASS\s*=\s*(\w+)(?:\s*\|\s*([\s\w]+\w))*")
  #  VALID_ARCHITECTURES           = IA32 X64 EBC
  REArchLine = re.compile(r"^\s*#\s*VALID_ARCHITECTURES\s*=\s*([\(\)\s\w]+[\w\)])*")
  def __init__(self, path, pkg):
    self.Path = path
    self.Pkg = pkg
  
  def IsLibraryFile(self):
    with open(self.Path, 'r') as f:
      line = f.readlines()
      FindLibraryClassLine = False
      self.Arch = ['ALL']
      for each in line:
        m = re.match(InfLibraryFile.RELibraryLine,each)
        if m != None:
          FindLibraryClassLine = True
          self.Module_Type_list = m.group(2)
          if self.Module_Type_list == None:
            self.Module_Type_list = ['ALL']
          else:
            self.Module_Type_list = re.sub(re.compile(r"\s+"), ' ',self.Module_Type_list)
            self.Module_Type_list = self.Module_Type_list.split(' ')
            if 'HOST_APPLICATION' in self.Module_Type_list:
              if len(self.Module_Type_list) == 1:
                return False
              self.Module_Type_list.remove('HOST_APPLICATION')
          self.LibraryClass = m.group(1)
        
        m = re.match(InfLibraryFile.REArchLine,each)
        if m != None:
          self.Arch = m.group(1)
          self.Arch = re.sub(re.compile(r'\([\w\s]*\)'), '', self.Arch)
          self.Arch = re.sub(re.compile(r"\s+"), ' ', self.Arch)
          self.Arch = self.Arch.split(' ')
          
      return FindLibraryClassLine

Dir =r"../"

Inf_file = []
for root, dirs, files in os.walk(Dir):
  for filename in files:
    if filename.endswith('.inf'):
      Pkg = root.split('\\')[1]
      inf = InfLibraryFile(os.path.join(root,filename), Pkg)
      if inf.IsLibraryFile():
        Inf_file.append(inf)




Sigle_Instance = []
Multi_Instance = []
Multi_Sigle_Instance = []
Multi_Multi_Instance = []

for one_library_instance in Inf_file:
  Same_name_library_instance = []
  repeate_count =0
  Module_Type = []
  Arch_List = []
  for another_library_instance in Inf_file:
    if one_library_instance.LibraryClass == another_library_instance.LibraryClass:
      repeate_count += 1
      Same_name_library_instance.append(another_library_instance)
      Module_Type += another_library_instance.Module_Type_list
      Arch_List += another_library_instance.Arch
  if repeate_count == 1:
    Sigle_Instance.append(one_library_instance)
  if repeate_count > 1:
    Multi_Instance.append(one_library_instance)

    if ('ALL' in Module_Type) and ('ALL' in Arch_List):
      Multi_Multi_Instance.append(one_library_instance)
      continue
    if 'ALL' in Module_Type:
      Is_Multiple_single = True
      for arch in Arch_List:
        Same_arch_Module_Type_List = []
        for One_same_libarary in Same_name_library_instance:
          if arch in One_same_libarary.Arch:
            Same_arch_Module_Type_List += One_same_libarary.Module_Type_list
        if 'ALL' in Same_arch_Module_Type_List:
          if len(Same_arch_Module_Type_List) != 1:
            Multi_Multi_Instance.append(one_library_instance)
            Is_Multiple_single = False
            break
        if len(set(Same_arch_Module_Type_List)) != len(Same_arch_Module_Type_List):
          Multi_Multi_Instance.append(one_library_instance)
          Is_Multiple_single = False
          break
      if not Is_Multiple_single:
        continue
      else:
        Multi_Sigle_Instance.append(one_library_instance)
    else:
      Is_Multiple_single = True
      for module in Module_Type:
        same_module_list=[]
        same_moudle_Arch_list = []
        for same_library_instance in Same_name_library_instance:
          if module in same_library_instance.Module_Type_list:
            same_moudle_Arch_list += same_library_instance.Arch
        if 'ALL' in same_moudle_Arch_list:
          if len(same_moudle_Arch_list) != 1:
            Multi_Multi_Instance.append(one_library_instance)
            Is_Multiple_single = False
            break
        if len(set(same_moudle_Arch_list)) != len(same_moudle_Arch_list):
          Multi_Multi_Instance.append(one_library_instance)
          Is_Multiple_single = False
          break
      if not Is_Multiple_single:
        continue
      else:
        Multi_Sigle_Instance.append(one_library_instance)

    

      

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Font, Alignment

wb = Workbook()
sht_4 = wb.create_sheet("Sigle-Library-Instance")
sht_5 = wb.create_sheet("Multiple-Sigle-Library-Instance")
sht_6 = wb.create_sheet("Multiple-Library-Instance")

Sigle_Instance_sorted = sorted(Sigle_Instance, key=lambda k: k.Pkg)
for e in Sigle_Instance_sorted:
  if e.Path.find("edk2-platforms") == -1:
    sht_4.append([e.Pkg, e.LibraryClass, ' '.join(e.Module_Type_list), ' '.join(e.Arch), e.Path[2:]])

Multi_Sigle_Instance_sorted = sorted(Multi_Sigle_Instance, key=lambda k: k.LibraryClass)
for e in Multi_Sigle_Instance_sorted:
  if e.Path.find("edk2-platforms") == -1:
    sht_5.append([e.Pkg, e.LibraryClass, ' '.join(e.Module_Type_list), ' '.join(e.Arch), e.Path[2:]])

Multi_Multi_Instance_sorted = sorted(Multi_Multi_Instance, key=lambda k: k.LibraryClass)
for e in Multi_Multi_Instance_sorted:
  if e.Path.find("edk2-platforms") == -1:
    sht_6.append([e.Pkg, e.LibraryClass, ' '.join(e.Module_Type_list), ' '.join(e.Arch), e.Path[2:]])

wb.save('Edk2 Library.xlsx')

Include_inf = Sigle_Instance + Multi_Sigle_Instance
Include_Module_ALL = set()
Include_Arch_ALL = set()
Include_Package_ALL = set()

for e in Include_inf:
  if ('IA32' in e.Arch) and ('X64' in e.Arch):
    e.Arch = ['common']
  if ('ALL' in e.Arch) or ('ANY' in e.Arch):
    e.Arch = ['common']
for e in Include_inf:
  Include_Module_ALL = Include_Module_ALL | set(e.Module_Type_list)
  Include_Arch_ALL = Include_Arch_ALL | set(e.Arch)
  if e.Path.find("edk2-platforms") == -1:
    Include_Package_ALL.add(e.Pkg)

for pkg in Include_Package_ALL:
  Modified = False
  if pkg == 'NetworkPkg' or pkg == 'RedfishPkg':
    continue
  pkg_Include_inf = []
  for e in Include_inf:
    if e.Pkg == pkg:
      pkg_Include_inf.append(e)
  with open(os.path.join(Dir, r'edk2/{0}/{1}Libs.dsc.inc'.format(pkg,pkg[:-3])), 'w') as f:
    # find common.all
    for arch in Include_Arch_ALL:
      for module in Include_Module_ALL:
        write_file = []
        for e in pkg_Include_inf:
          if (arch in e.Arch) and (module in e.Module_Type_list):
            write_file.append(e)          
        if len(write_file) != 0:
          if module == 'ALL':
            f.write('\n[LibraryClasses.{0}]\n'.format(arch))
          else:
            f.write('\n[LibraryClasses.{0}.{1}]\n'.format(arch, module))
          for e in write_file:
            path = e.Path
            path = path[path.find('edk2')+5:]
            f.write('  {0}|{1}\n'.format(e.LibraryClass,path))
            Modified = True

  if Modified:
    mydir =  os.getcwd()
    os.chdir(os.path.join(Dir,'edk2'))
    ScmCmd = 'git add *'
    RunCommand(ScmCmd)
    ScmCmd = 'git commit -s -m "{0}:Add {1}Libs.dsc.inc \n"'.format(pkg, pkg[:-3])
    RunCommand(ScmCmd)
    os.chdir(mydir)
  



